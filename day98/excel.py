from openpyxl import load_workbook

wb = load_workbook('test.xlsx')
ws = wb.active


# 获取整张表的数据，返回一个列表data_list
def data2list():
    data_list = []
    for row in list(ws.iter_rows()):
        row_data_list = []
        for i in row:
            row_data_list.append(i.value)
            # print(row_data_list)
        data_list.append(row_data_list)
    # print(data_list)
    return data_list


# 将匹配到的数据筛选出来，返回一个列表match_list
def match_res(str, row, data_list):
    match_list = []
    match_list.append(data_list[row])
    print(data_list[row])
    for row_data_list in data_list:
        if row_data_list[row] == str:
            match_list.append(row_data_list)
    return match_list


# 获取分类
def kind_list(data_list):
    kind = []
    for item in data_list:
        kind.append(item[0])

    kind = sorted(list(set(kind[1:])), reverse=True)
    return kind


data_list = data2list()
kind = kind_list(data_list)

for item in kind:
    match_list = match_res(item, 0, data_list)

    if item in wb.sheetnames:
        ws2 = wb[item]
    else:
        ws2 = wb.create_sheet(item, 1)
    # for i in range(len(match_list)):
    #     # print(i)
    #     for j in range(len(match_list[i])):
    #         ws2.cell(i + 1, j + 1, match_list[i][j])
    for item in match_list:
        ws2.append(item)

wb.save('test2.xlsx')
