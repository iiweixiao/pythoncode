import os

from openpyxl import load_workbook

wb = load_workbook('1.xlsx')

# # print(wb.sheetnames)
# s1 = wb.worksheets[0]
# cell = s1.cell(1, 3)
# print(cell.value)

# for name in wb.sheetnames:
#     sheet = wb[name]
#     a = 0
#     for row in sheet.iter_rows(min_row=2):
#         cell = row[1]
#         a = a + cell.value
#     print(sheet.title, a)

index = 1
for s in wb.worksheets:
    a = 0
    for row in s.iter_rows(min_row=2, max_row=4):
        cell = row[0]

        a = a + cell.value
        # print(index, cell.border)
        # index += 1
        # print(type(cell.value))
    print(s.title, a)


# folader_path = os.path.join('file_data')
# for file_name in os.listdir(folader_path):
#     filepath = os.path.join(folader_path, file_name)
#     wb = load_workbook(filepath)
#
# data_dict = {}
# for name in wb.sheetnames:
#     sheet_x1 = wb[name]
#     a = 0
#     NewName = file_name.split(".")[0]
#     data_dict[NewName] = {}
