from openpyxl import Workbook

wb = Workbook()
ws1 = wb.active

ws2 = wb.create_sheet('sheet2', 1)
ws3 = wb.create_sheet('sheet3', 2)

wb.move_sheet(ws3, -1)


for cells in ws1['a1:c4']:
    for cell in cells:
        print(cell)
"""
<Cell 'Sheet'.A1>
<Cell 'Sheet'.B1>
<Cell 'Sheet'.C1>
<Cell 'Sheet'.A2>
<Cell 'Sheet'.B2>
<Cell 'Sheet'.C2>
<Cell 'Sheet'.A3>
<Cell 'Sheet'.B3>
<Cell 'Sheet'.C3>
<Cell 'Sheet'.A4>
<Cell 'Sheet'.B4>
<Cell 'Sheet'.C4>
"""

print(wb.sheetnames)