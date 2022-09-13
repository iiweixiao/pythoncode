# 1.打升excel

from openpyxl import workbook

# 创建excel 同时创建一个新的sheet
wb = workbook.Workbook()

sheet = wb.worksheets[0]

sheet.cell(1, 1).value = '这是我的第一个单元格！'

wb.save('test.xlsx')

# 2.原有的excel

from openpyxl import load_workbook

wb = load_workbook('test.xlsx')
sheet = wb.worksheets[0]

sheet.cell(1, 1).value = '修改我的第一个单元格'
wb.save('test.xlsx')