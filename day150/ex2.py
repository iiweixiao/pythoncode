import os
from openpyxl import load_workbook, workbook

user = input("用户名：")

password = input("密码：")

# 1.excel存在原有的文件里面去气

# 2．不存在这个文件，新建一个

filepath = os.path.join('dv', 'demo.xlsx')
if os.path.exists(filepath):
    wb = load_workbook(filepath)
    sheet = wb.worksheets[0]
    next_row_index = sheet.max_row + 1
else:
    wb = workbook.Workbook()
    sheet = wb.worksheets[0]
    next_row_index = 1

sheet.cell(next_row_index, 1).value = user
sheet.cell(next_row_index, 2).value = password

wb.save(filepath)
