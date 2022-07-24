from openpyxl import Workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import FORMULAE

wb = Workbook()
ws = wb.active


# print(len(FORMULAE))


def is_exist(str):
    if str.upper() in FORMULAE:
        print('True')
    else:
        print('False')


is_exist('vlookup')

ws.append(["价格1", "价格2", "求和", "平均值"])
ws.append([54, 21])
ws.append([12, 64])
ws.append([72, 14])

# 求和
ws["c2"] = "=SUM(A2:B2)"
# 求平均值
ws["d2"] = "=AVERAGE(A2:B2)"

# 复制公式到其他单元格
# ws["c3"] = Translator(formula="=SUM(A2:B2)", origin="c2").translate_formula("c3")
# ws["c4"] = Translator(formula="=SUM(A2:B2)", origin="c2").translate_formula("c4")

# ws["d3"] = Translator(formula="=average(A2:B2)", origin="d2").translate_formula("d3")
# ws["d4"] = Translator(formula="=average(A2:B2)", origin="d2").translate_formula("d4")

# 用for循环复制公式
for cell in ws["c3:c4"]:
    cell[0].value = Translator(formula="=SUM(A2:B2)", origin="c2").translate_formula(cell[0].coordinate)
for cell in ws["d3:d4"]:
    cell[0].value = Translator(formula="=average(A2:B2)", origin="d2").translate_formula(cell[0].coordinate)

wb.save("e2.xlsx")
