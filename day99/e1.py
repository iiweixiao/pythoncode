from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = 'abc'

x = 1
for i in range(1, 11):
    for j in range(1, 6):
        ws.cell(i, j, x)
        x += 1

for cells in ws['a1:c4']:
    for cell in cells:
        print(cell.value)

print('---------')

# 同ws['b1:e3']，第1-3行，第2-5列
for rows in ws.iter_rows(1, 3, 2, 5):
    for row in rows:
        print(row.value)

# 合并单元格，取消合并
ws.merge_cells('b2:c3')
ws.unmerge_cells('b2:c3')

# 插入行/列，在第2行/列前插入1行/列
ws.insert_rows(2, 1)
ws.insert_cols(2, 1)

# 删除行/列
ws.delete_rows(2, 1)
ws.delete_cols(2, 1)

# 目标单元格向下移动10行
ws.move_range('a1:b2', 10, 0)

wb.save('e2.xlsx')